import csv
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_CASE_FIELDS = ["user_judgment", "user_reason"]
OPTIONAL_CASE_FIELDS = ["gid", "title", "content"]


def _detect_delimiter(file_path):
    if file_path.suffix.lower() == ".xlsx":
        return None

    if file_path.suffix.lower() == ".tsv":
        return "\t"

    sample = file_path.read_text(encoding="utf-8-sig")[:2048]
    try:
        return csv.Sniffer().sniff(sample).delimiter
    except csv.Error:
        return ","


def read_case_columns(file_path):
    file_path = Path(file_path)
    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        try:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            return [str(value).strip() for value in first_row if value is not None]
        finally:
            workbook.close()

    delimiter = _detect_delimiter(file_path)
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        return next(reader, [])


def _load_case_studies_from_excel(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active

    try:
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise ValueError(f"案例表格为空或缺少表头：{file_path}")

    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    return _normalize_case_rows(file_path, header, rows[1:])


def _normalize_case_rows(file_path, fieldnames, rows):
    missing_fields = [field for field in REQUIRED_CASE_FIELDS if field not in fieldnames]
    if missing_fields:
        raise ValueError(f"案例表格缺少必填列 {missing_fields}：{file_path}")

    has_gid = "gid" in fieldnames
    has_legacy_content = all(field in fieldnames for field in ["title", "content"])
    if not has_gid and not has_legacy_content:
        raise ValueError(
            f"案例表格至少需要 gid 列，或同时包含 title/content 列：{file_path}"
        )

    case_studies = []
    for row in rows:
        if isinstance(row, dict):
            cleaned_row = {
                key.strip(): (value or "").strip()
                for key, value in row.items()
                if key
            }
        else:
            cleaned_row = {}
            for index, key in enumerate(fieldnames):
                if not key:
                    continue
                value = row[index] if index < len(row) else ""
                cleaned_row[key] = str(value).strip() if value is not None else ""

        if not any(cleaned_row.values()):
            continue

        normalized = {field: cleaned_row.get(field, "") for field in OPTIONAL_CASE_FIELDS}
        normalized.update({
            "user_judgment": cleaned_row.get("user_judgment", ""),
            "user_reason": cleaned_row.get("user_reason", ""),
            "raw": cleaned_row,
        })
        case_studies.append(normalized)

    return case_studies


def load_case_studies(file_path):
    file_path = Path(file_path)
    if file_path.suffix.lower() == ".xlsx":
        return _load_case_studies_from_excel(file_path)

    delimiter = _detect_delimiter(file_path)
    with file_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)

        if not reader.fieldnames:
            raise ValueError(f"案例表格为空或缺少表头：{file_path}")

        reader.fieldnames = [field.strip() if field else "" for field in reader.fieldnames]
        return _normalize_case_rows(file_path, reader.fieldnames, reader)
